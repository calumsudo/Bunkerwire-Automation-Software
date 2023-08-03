import PyPDF2
import re
from datetime import datetime
from pdf_extractors import extract_mfhod00_mid, extract_aagpd00_mid, extract_aawwx00_mid, extract_mfrdd00_mid, extract_puafn00_mid, extract_aartg00_mid, extract_mfgbd00_mid, extract_aakab00_mid, extract_aarsu00_mid, extract_mffjd00_mid, extract_puaxp00_mid, extract_aarkh00_mid, extract_aaxyp00_mid, extract_mfspd00_mid, extract_puaft00_mid, extract_aalmz00_mid, extract_aaxyo00_mid, extract_mfskd00_mid, extract_puafr00_mid, extract_aavbn00_mid, extract_aaxys00_mid
import tkinter as tk
from tkinter import filedialog
import openpyxl  # Import the openpyxl library

# Create the main window
root = tk.Tk()
root.title("SPI Bunkerwire Automation")



#Extract and Format Date Function Completed
def extract_and_format_date(pdf_file_path):
    with open(pdf_file_path, 'rb') as file:
        reader = PyPDF2.PdfReader(file)
        num_pages = len(reader.pages)

        # Extract the entire text from all pages
        pdf_text = ""
        for page_num in range(num_pages):
            pdf_text += reader.pages[page_num].extract_text()

        # Search for the date pattern (e.g., "July 14, 2023") in the text
        date_pattern = r'[A-Za-z]+\s+\d{1,2},\s+\d{4}'
        date_matches = re.findall(date_pattern, pdf_text)

        # Convert the date matches to the desired format
        formatted_dates = []
        for date_str in date_matches:
            date_obj = datetime.strptime(date_str, "%B %d, %Y")
            formatted_date = date_obj.strftime("%m/%d/%Y %I:%M:%S %p")
            formatted_dates.append(formatted_date)

        # Return the first formatted date or None if there are no date matches
        return formatted_dates[0] if formatted_dates else None

#Display the Data in the terminal for testing purposes
def extract_and_display_data(pdf_file_path):

    date = extract_and_format_date(pdf_file_path)

    data_values = {
        'DATE': date,
        'HOU VLSFO 0.5%': extract_mfhod00_mid(pdf_file_path),
        'HOU IFO380 3.5%': extract_aagpd00_mid(pdf_file_path),
        'HOU MGO 0.1%': extract_aawwx00_mid(pdf_file_path),
        'RDAM VLSFO 0.5%': extract_mfrdd00_mid(pdf_file_path),
        'RDAM IFO380 3.5%': extract_puafn00_mid(pdf_file_path),
        'RDAM MGO 0.1%': extract_aartg00_mid(pdf_file_path),
        'GIB VLSFO 0.5%': extract_mfgbd00_mid(pdf_file_path),
        'GIB IFO380 3.5%': extract_aakab00_mid(pdf_file_path),
        'GIB MGO 0.1%': extract_aarsu00_mid(pdf_file_path),
        'FUJAIRAH VLSFO 0.5%': extract_mffjd00_mid(pdf_file_path),
        'FUJAIRAH IFO380 3.5%': extract_puaxp00_mid(pdf_file_path),
        'FUJAIRAH MGO 0.5%': extract_aarkh00_mid(pdf_file_path),
        'FUJAIRAH MGO 0.1%': extract_aaxyp00_mid(pdf_file_path),
        'SPORE VLSFO 0.5%': extract_mfspd00_mid(pdf_file_path),
        'SPORE IFO380 3.5%': extract_puaft00_mid(pdf_file_path),
        'SPORE MGO 0.5%': extract_aalmz00_mid(pdf_file_path),
        'SPORE MGO 0.1%': extract_aaxyo00_mid(pdf_file_path),
        'ULSAN VLSFO 0.5%': extract_mfskd00_mid(pdf_file_path),
        'ULSAN IFO380 3.5%': extract_puafr00_mid(pdf_file_path),
        'ULSAN MGO 0.5%': extract_aavbn00_mid(pdf_file_path),
        'ULSAN MGO 0.1% ': extract_aaxys00_mid(pdf_file_path),
    }

    return data_values


def browse_pdf():
    file_path = filedialog.askopenfilename(filetypes=[("PDF Files", "*.pdf")])
    pdf_entry.delete(0, tk.END)
    pdf_entry.insert(tk.END, file_path)

def browse_excel():
    file_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xls *.xlsx")])
    excel_entry.delete(0, tk.END)
    excel_entry.insert(tk.END, file_path)


def print_message(message):
    messages_label.config(text=messages_label.cget('text') + message + "\n")

def submit_files():
    pdf_file = pdf_entry.get()
    excel_file = excel_entry.get()

    if pdf_file:
        # Extract and format the date
        date = extract_and_format_date(pdf_file)

        # Extract and display data
        data_values = extract_and_display_data(pdf_file)

        if excel_file:
            # Load the existing Excel file
            wb = openpyxl.load_workbook(excel_file)
            sheet = wb.active

            # Find the first empty row to append the data
            row = sheet.max_row + 1

            # Write the data to the corresponding columns
            for col_name, value in data_values.items():
                # Search for the column with the matching column name
                for col_index, cell in enumerate(sheet[1], 1):
                    if cell.value == col_name:
                        col_letter = openpyxl.utils.get_column_letter(col_index)
                        sheet[f"{col_letter}{row}"] = value
                        break

            # Save the Excel file
            wb.save(excel_file)

            print_message("Data appended to the Excel file.")
        else:
            print_message("Please upload an Excel file.")
    else:
        print_message("Please upload a PDF file.")


# Create frames to organize the sections
pdf_frame = tk.Frame(root)
pdf_frame.pack(pady=10)
excel_frame = tk.Frame(root)
excel_frame.pack(pady=10)

# PDF section
pdf_label = tk.Label(pdf_frame, text="Upload PDF:")
pdf_label.pack(side=tk.LEFT)
pdf_entry = tk.Entry(pdf_frame, width=50)
pdf_entry.pack(side=tk.LEFT)
pdf_button = tk.Button(pdf_frame, text="Browse", command=browse_pdf)
pdf_button.pack(side=tk.LEFT)

# Excel section
excel_label = tk.Label(excel_frame, text="Upload Excel:")
excel_label.pack(side=tk.LEFT)
excel_entry = tk.Entry(excel_frame, width=50)
excel_entry.pack(side=tk.LEFT)
excel_button = tk.Button(excel_frame, text="Browse", command=browse_excel)
excel_button.pack(side=tk.LEFT)

# Submit button
submit_button = tk.Button(root, text="Submit", command=submit_files)
submit_button.pack(pady=20)


# Create a Text widget to display messages
# Create a Label widget to display messages
messages_label = tk.Label(root, height=5, width=70, wraplength=500, anchor='w', justify='center')
messages_label.pack(pady=1)

# Start the main event loop
root.mainloop()
